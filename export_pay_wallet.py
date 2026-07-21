#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok Seller — LỊCH SỬ RÚT TIỀN (ví/earnings-withdrawal) qua EXPORT.

Chạy local:
  python export_pay_wallet.py --last-days 3 --cookie cookie.txt --shop-name RHYSMAN
  python export_pay_wallet.py --month 2026-06 --cookie cookie_hair.txt --shop-name HAIR

Chạy GitHub Actions:
  - Truyền cookie qua biến môi trường TIKTOK_COOKIE.
  - Truyền service-account JSON qua GCP_SERVICE_ACCOUNT_JSON,
    hoặc dùng GOOGLE_APPLICATION_CREDENTIALS / google-github-actions/auth.

Mặc định nạp vào:
  rhysman-data-warehouse-488306.rhysman.fact_tiktok_pay_wallet

Idempotent theo từng shop:
  DELETE đúng seller_id + khoảng request_time, sau đó INSERT lại.
"""

import argparse
import datetime
import gzip
import io
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Optional

# Fallback dành cho chạy local. Trên GitHub Actions ưu tiên TIKTOK_COOKIE.
DEFAULT_COOKIE_FILE = "cookie.txt"
SELLER_ID: Optional[str] = None
SHOP_NAME = "UNKNOWN"

# BigQuery đích — có thể ghi đè bằng environment variables.
DEFAULT_BQ_PROJECT = "rhysman-data-warehouse-488306"
DEFAULT_BQ_DATASET = "rhysman"
DEFAULT_BQ_TABLE = "fact_tiktok_pay_wallet"

# Sheet 'Lịch sử rút tiền' — 7 cột. Tiền là VND (số nguyên, ÂM khi rút).
SHEET_NAME = "Lịch sử rút tiền"
EXPECTED_HEADERS_WD = [
    "Loại giao dịch",
    "ID tham chiếu",
    "Thời gian yêu cầu",
    "Số tiền",
    "Trạng thái",
    "Thời gian thành công",
    "Tài khoản ngân hàng",
]

# (out_name, src_header, kind). kind: text/date/num/const
OUT_SPEC = [
    ("transaction_type", "Loại giao dịch", "text"),
    ("reference_id", "ID tham chiếu", "text"),
    ("request_time", "Thời gian yêu cầu", "date"),
    ("amount", "Số tiền", "num"),
    ("status", "Trạng thái", "text"),
    ("success_time", "Thời gian thành công", "date"),
    ("bank_account", "Tài khoản ngân hàng", "text"),
    ("seller_id", None, "const"),
]
OUT_NAMES = [name for name, _, _ in OUT_SPEC]
DATE_OUT = {"request_time", "success_time"}

TZ = datetime.timezone(datetime.timedelta(hours=7))
BASE = "https://seller-vn.tiktok.com"
API = BASE + "/api/v2/pay/settlement/file"

HEADERS = {
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    "accept": "*/*",
    "origin": BASE,
    "referer": BASE + "/finance/transactions?shop_region=VN&tab=settled_tab",
    "x-tt-oec-region": "VN",
}
TIME_TYPE = {"settle": 1, "placed": 2}


def common_params() -> dict:
    if not SELLER_ID:
        raise RuntimeError("SELLER_ID chưa được khởi tạo.")
    return {
        "locale": "vi-VN",
        "language": "vi-VN",
        "oec_seller_id": SELLER_ID,
        "seller_id": SELLER_ID,
        "aid": "4068",
        "app_name": "i18n_ecom_shop",
        "device_platform": "web",
        "cookie_enabled": "true",
    }


def read_cookie(cookie_file: Optional[str], cookie_env: str) -> str:
    """Ưu tiên cookie từ environment; nếu không có thì đọc file local."""
    cookie = os.getenv(cookie_env, "").strip()
    if cookie:
        return cookie

    if cookie_file:
        path = Path(cookie_file)
        if path.is_file():
            cookie = path.read_text(encoding="utf-8").strip()
            if cookie:
                return cookie

    sys.exit(
        f"[CONFIG] Không có cookie. Hãy đặt secret vào biến {cookie_env} "
        f"hoặc truyền --cookie <đường_dẫn_file>."
    )


def extract_seller_id(cookie: str) -> str:
    """Lấy seller ID từ cookie TikTok Seller Center."""
    patterns = [
        r"(?:^|;\s*)oec_seller_id_unified_seller_env=(\d+)",
        r"(?:^|;\s*)oec_seller_id=(\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, cookie)
        if match:
            return match.group(1)
    sys.exit(
        "[CONFIG] Không thấy oec_seller_id_unified_seller_env trong cookie. "
        "Cookie có thể sai, thiếu hoặc đã hết hạn."
    )


def req(url: str, cookie: str, body=None, raw: bool = False, tries: int = 5):
    for attempt in range(tries):
        try:
            data = json.dumps(body).encode("utf-8") if body is not None else None
            headers = {**HEADERS, "cookie": cookie}
            if body is not None:
                headers["content-type"] = "application/json"

            request = urllib.request.Request(
                url,
                data=data,
                method="POST" if body is not None else "GET",
                headers=headers,
            )
            with urllib.request.urlopen(request, timeout=60) as response:
                buf = response.read()
                if response.headers.get("content-encoding") == "gzip":
                    buf = gzip.decompress(buf)

            if raw:
                return buf

            try:
                return json.loads(buf)
            except json.JSONDecodeError as exc:
                preview = buf[:300].decode("utf-8", errors="replace")
                raise RuntimeError(f"API trả về dữ liệu không phải JSON: {preview}") from exc

        except Exception as exc:
            if attempt == tries - 1:
                raise
            print(
                f"    retry {attempt + 1}/{tries - 1} ({type(exc).__name__}: {exc})",
                flush=True,
            )
            time.sleep(2 * (attempt + 1))


def list_files(cookie: str) -> list:
    url = f"{API}/list?" + urllib.parse.urlencode(
        {**common_params(), "page": 1, "page_size": 20, "size": 20, "from": 0}
    )
    response = req(url, cookie)
    if response.get("code") not in (None, 0):
        raise RuntimeError(
            f"Không lấy được danh sách export: {response.get('code')} {response.get('message')}"
        )
    return (response.get("data") or {}).get("files", [])


def _num(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return int(number) if number.is_integer() else number

    text = str(value).strip().replace(" ", "").replace(",", "")
    if not text or text == "-":
        return 0
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return 0


def _to_date(value) -> Optional[str]:
    """Chuẩn hóa date/datetime/chuỗi thành YYYY-MM-DD cho BigQuery DATE."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()

    text = str(value).strip()
    if not text or text in {"-", "/"}:
        return None

    # Loại bỏ phần giờ nếu Excel/API trả về dạng YYYY-MM-DD HH:MM:SS.
    text = text.replace("/", "-").split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    raise ValueError(f"Không đọc được giá trị ngày: {value!r}")


def slim_wallet(src, begin_date: str, end_date: str) -> list[dict]:
    """Đọc sheet lịch sử rút tiền và lọc đúng kỳ yêu cầu."""
    import openpyxl

    if isinstance(src, (bytes, bytearray)):
        src = io.BytesIO(src)

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"[WALLET] Không thấy sheet '{SHEET_NAME}' trong file export.")

    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        sys.exit(f"[WALLET] Sheet '{SHEET_NAME}' đang trống.")

    idx = {str(item).strip(): i for i, item in enumerate(header) if item is not None}

    # Lá chắn schema: TikTok đổi cột thì dừng, không âm thầm nạp sai.
    actual = [str(item).strip() for item in header if item is not None]
    removed = [name for name in EXPECTED_HEADERS_WD if name not in idx]
    added = [name for name in actual if name not in EXPECTED_HEADERS_WD]
    if removed or added:
        message = ["[WALLET] CẤU TRÚC SHEET THAY ĐỔI — DỪNG để kiểm tra:"]
        if removed:
            message.append(f"  • Mất/đổi tên cột: {removed}")
        if added:
            message.append(f"  • Cột mới xuất hiện: {added}")
        message.append("  -> Cập nhật EXPECTED_HEADERS_WD / OUT_SPEC rồi chạy lại.")
        sys.exit("\n".join(message))

    plan = [
        (name, kind, idx[src_header] if src_header is not None else None)
        for name, src_header, kind in OUT_SPEC
    ]

    reference_index = idx["ID tham chiếu"]
    records = []
    outside_count = 0

    for row in rows:
        if not row or row[reference_index] in (None, ""):
            continue

        record = {}
        for name, kind, position in plan:
            if kind == "num":
                record[name] = _num(row[position])
            elif kind == "date":
                record[name] = _to_date(row[position])
            elif kind == "const":
                record[name] = SELLER_ID
            else:
                record[name] = (
                    str(row[position]).strip() if row[position] not in (None, "") else None
                )

        request_date = record["request_time"]
        if request_date is not None and begin_date <= request_date <= end_date:
            records.append(record)
        else:
            outside_count += 1

    earnings_count = sum(
        1 for record in records if record["transaction_type"] == "Earnings"
    )
    withdrawal_count = sum(
        1 for record in records if record["transaction_type"] == "Withdrawal"
    )
    print(
        f"[WALLET][{SHOP_NAME}] {len(records)} dòng trong kỳ "
        f"({earnings_count} Earnings, {withdrawal_count} Withdrawal); "
        f"bỏ {outside_count} dòng ngoài kỳ.",
        flush=True,
    )
    return records


def write_records_xlsx(records: list[dict], out_path: str) -> None:
    import openpyxl

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("wallet")
    ws.append(OUT_NAMES)
    for record in records:
        ws.append([record[name] for name in OUT_NAMES])
    wb.save(out_path)
    print(f"[XLSX][{SHOP_NAME}] -> {out_path}", flush=True)


def create_bq_client():
    """Hỗ trợ cả GitHub secret JSON, ADC và GOOGLE_APPLICATION_CREDENTIALS."""
    from google.cloud import bigquery

    project_id = os.getenv("GCP_PROJECT_ID", DEFAULT_BQ_PROJECT).strip()
    credentials_json = os.getenv("GCP_SERVICE_ACCOUNT_JSON", "").strip()

    if credentials_json:
        from google.oauth2 import service_account

        try:
            info = json.loads(credentials_json)
        except json.JSONDecodeError as exc:
            raise RuntimeError("GCP_SERVICE_ACCOUNT_JSON không phải JSON hợp lệ.") from exc
        credentials = service_account.Credentials.from_service_account_info(info)
        project_id = info.get("project_id") or project_id
        return bigquery.Client(credentials=credentials, project=project_id)

    # ADC: dùng được với GOOGLE_APPLICATION_CREDENTIALS hoặc google-github-actions/auth.
    return bigquery.Client(project=project_id)


def load_to_bq(records: list[dict], begin_date: str, end_date: str) -> None:
    """DELETE + INSERT theo seller_id để hai shop không xóa dữ liệu của nhau."""
    from google.cloud import bigquery

    client = create_bq_client()
    dataset = os.getenv("BQ_DATASET", DEFAULT_BQ_DATASET).strip()
    table = os.getenv("BQ_TABLE", DEFAULT_BQ_TABLE).strip()
    full_table = f"{client.project}.{dataset}.{table}"

    def bq_type(name: str) -> str:
        if name in {
            "transaction_type",
            "reference_id",
            "seller_id",
            "status",
            "bank_account",
        }:
            return "STRING"
        if name in DATE_OUT:
            return "DATE"
        return "INT64"

    schema = [bigquery.SchemaField(name, bq_type(name)) for name in OUT_NAMES]
    columns_ddl = ", ".join(f"`{name}` {bq_type(name)}" for name in OUT_NAMES)
    client.query(f"CREATE TABLE IF NOT EXISTS `{full_table}` ({columns_ddl})").result()

    delete_sql = f"""
        DELETE FROM `{full_table}`
        WHERE seller_id = @seller_id
          AND request_time BETWEEN @begin_date AND @end_date
    """
    query_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("seller_id", "STRING", SELLER_ID),
            bigquery.ScalarQueryParameter("begin_date", "DATE", begin_date),
            bigquery.ScalarQueryParameter("end_date", "DATE", end_date),
        ]
    )
    client.query(delete_sql, job_config=query_config).result()
    print(
        f"[BQ][{SHOP_NAME}] Đã xóa dữ liệu cũ của seller {SELLER_ID} "
        f"trong {begin_date}..{end_date}",
        flush=True,
    )

    load_config = bigquery.LoadJobConfig(
        schema=schema,
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
    )
    client.load_table_from_json(records, full_table, job_config=load_config).result()
    print(
        f"[BQ][{SHOP_NAME}] Đã nạp {len(records)} dòng -> {full_table}",
        flush=True,
    )


def parse_period(args) -> tuple[int, int, str]:
    selected_modes = sum(
        [
            bool(args.month),
            bool(args.d_from or args.d_to),
            args.last_days is not None,
        ]
    )
    if selected_modes > 1:
        sys.exit("[CONFIG] Chỉ được chọn một mode: --month, --from/--to hoặc --last-days.")

    if bool(args.d_from) != bool(args.d_to):
        sys.exit("[CONFIG] Phải truyền đồng thời cả --from và --to.")

    if args.month:
        try:
            year, month = map(int, args.month.split("-"))
            first_day = datetime.datetime(year, month, 1, tzinfo=TZ)
        except (ValueError, TypeError):
            sys.exit("[CONFIG] --month phải có định dạng YYYY-MM.")
        next_month = datetime.datetime(
            year + (month == 12), (month % 12) + 1, 1, tzinfo=TZ
        )
        begin = int(first_day.timestamp())
        end = int(next_month.timestamp()) - 1
        tag = args.month.replace("-", "_")
        return begin, end, tag

    if args.d_from and args.d_to:
        try:
            from_date = datetime.datetime.strptime(args.d_from, "%Y-%m-%d").date()
            to_date = datetime.datetime.strptime(args.d_to, "%Y-%m-%d").date()
        except ValueError:
            sys.exit("[CONFIG] --from và --to phải có định dạng YYYY-MM-DD.")
        if from_date > to_date:
            sys.exit("[CONFIG] --from không được lớn hơn --to.")

        begin_dt = datetime.datetime.combine(from_date, datetime.time.min, tzinfo=TZ)
        end_dt = datetime.datetime.combine(to_date, datetime.time.max, tzinfo=TZ)
        return int(begin_dt.timestamp()), int(end_dt.timestamp()), f"{args.d_from}_{args.d_to}"

    days = args.last_days if args.last_days is not None else 2
    if days <= 0:
        sys.exit("[CONFIG] --last-days phải lớn hơn 0.")

    today = datetime.datetime.now(TZ).date()
    begin_date = today - datetime.timedelta(days=days - 1)
    begin_dt = datetime.datetime.combine(begin_date, datetime.time.min, tzinfo=TZ)
    end_dt = datetime.datetime.combine(today, datetime.time.max, tzinfo=TZ)
    return int(begin_dt.timestamp()), int(end_dt.timestamp()), f"last{days}d_{today:%Y%m%d}"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="TikTok Seller wallet (earnings/withdrawal) export"
    )
    parser.add_argument("--month", help="YYYY-MM")
    parser.add_argument("--from", dest="d_from", help="YYYY-MM-DD")
    parser.add_argument("--to", dest="d_to", help="YYYY-MM-DD, bao gồm cả ngày này")
    parser.add_argument(
        "--last-days",
        type=int,
        default=None,
        help="Kéo N ngày gần nhất; mặc định 2 nếu không chọn mode khác",
    )
    parser.add_argument("--by", choices=["settle", "placed"], default="settle")
    parser.add_argument(
        "--cookie",
        default=DEFAULT_COOKIE_FILE,
        help="File cookie dùng khi chạy local; GitHub ưu tiên environment",
    )
    parser.add_argument(
        "--cookie-env",
        default="TIKTOK_COOKIE",
        help="Tên biến môi trường chứa cookie; mặc định TIKTOK_COOKIE",
    )
    parser.add_argument(
        "--shop-name",
        default=os.getenv("SHOP_NAME", "UNKNOWN"),
        help="Tên shop chỉ dùng cho log, ví dụ RHYSMAN hoặc HAIR",
    )
    parser.add_argument("--xlsx", help="Tùy chọn: ghi thêm file XLSX đã làm gọn")
    parser.add_argument("--no-bq", action="store_true", help="Không nạp BigQuery")
    parser.add_argument("--keep-raw", action="store_true", help="Giữ file export đầy đủ")
    parser.add_argument(
        "--wait",
        type=int,
        default=240,
        help="Số giây tối đa chờ TikTok tạo file; mặc định 240",
    )
    args = parser.parse_args()

    global SELLER_ID, SHOP_NAME
    SHOP_NAME = args.shop_name.strip().upper() or "UNKNOWN"
    cookie = read_cookie(args.cookie, args.cookie_env)
    SELLER_ID = extract_seller_id(cookie)
    print(
        f"[CONFIG][{SHOP_NAME}] seller_id={SELLER_ID} | cookie_source={args.cookie_env}",
        flush=True,
    )

    begin, end, tag = parse_period(args)
    begin_date = datetime.datetime.fromtimestamp(begin, TZ).date().isoformat()
    end_date = datetime.datetime.fromtimestamp(end, TZ).date().isoformat()
    print(
        f"[EXPORT][{SHOP_NAME}] kỳ {begin_date} -> {end_date} | theo "
        f"{'ngày quyết toán' if args.by == 'settle' else 'ngày tạo đơn'}",
        flush=True,
    )

    # Chụp danh sách hiện tại để nhận ra file mới.
    before = {
        item.get("file_id")
        for item in list_files(cookie)
        if item.get("file_id") is not None
    }

    body = {
        "period": {
            "begin_date": str(begin),
            "end_date": str(end),
            "time_type": TIME_TYPE[args.by],
        },
        "file_type": 1,
        "statement_version": 0,
    }
    response = req(
        f"{API}/export?" + urllib.parse.urlencode(common_params()),
        cookie,
        body=body,
    )
    if response.get("code") != 0:
        sys.exit(
            f"[EXPORT][{SHOP_NAME}] Tạo job lỗi: "
            f"{response.get('code')} {response.get('message')}"
        )
    print(f"[EXPORT][{SHOP_NAME}] Đã tạo job, chờ server sinh file...", flush=True)

    started_at = time.time()
    file_id = None
    file_name = ""
    while time.time() - started_at < args.wait:
        time.sleep(3)
        for item in list_files(cookie):
            current_id = item.get("file_id")
            if current_id not in before and item.get("status") == 2:
                file_id = current_id
                file_name = item.get("file_name", "")
                break
        if file_id:
            break
        print(
            f"    [{SHOP_NAME}] đang xử lý ({int(time.time() - started_at)}s)",
            flush=True,
        )

    if not file_id:
        sys.exit(
            f"[EXPORT][{SHOP_NAME}] Quá {args.wait}s chưa thấy file mới. "
            "TikTok có thể đang chậm hoặc cookie đã hết hạn."
        )
    print(
        f"[EXPORT][{SHOP_NAME}] File sẵn sàng: {file_name} (id {file_id})",
        flush=True,
    )

    download_response = req(
        f"{API}/download?"
        + urllib.parse.urlencode({**common_params(), "file_id": file_id}),
        cookie,
    )
    relative_url = (download_response.get("data") or {}).get("url")
    if not relative_url:
        sys.exit(f"[EXPORT][{SHOP_NAME}] Không lấy được link tải.")

    download_url = urllib.parse.urljoin(BASE, relative_url)
    blob = req(download_url, cookie, raw=True)
    print(
        f"[EXPORT][{SHOP_NAME}] Tải xong file: {len(blob) // 1024} KB",
        flush=True,
    )

    if args.keep_raw:
        raw_path = f"_full_wallet_{SHOP_NAME.lower()}_{tag}.xlsx"
        Path(raw_path).write_bytes(blob)
        print(f"[EXPORT][{SHOP_NAME}] Giữ file đầy đủ: {raw_path}", flush=True)

    records = slim_wallet(blob, begin_date, end_date)
    if not records:
        print(
            f"[WALLET][{SHOP_NAME}] Không có dòng nào trong kỳ — không nạp BigQuery.",
            flush=True,
        )
        return

    if args.xlsx:
        write_records_xlsx(records, args.xlsx)

    if args.no_bq:
        print(f"[BQ][{SHOP_NAME}] Bỏ qua nạp BigQuery (--no-bq).", flush=True)
    else:
        load_to_bq(records, begin_date, end_date)


if __name__ == "__main__":
    main()
