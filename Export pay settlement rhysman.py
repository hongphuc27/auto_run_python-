#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok Seller — BÁO CÁO QUA EXPORT, end-to-end (nhanh nhất: cả kỳ trong 1 lần).
Luồng: tạo job export -> chờ server sinh file -> tải full 65 cột -> rút còn 21 cột -> xoá full.

VÍ DỤ:
  py -u export_pay_settlement.py                     # MẶC ĐỊNH: 2 ngày gần nhất -> nạp BQ (hợp cron)
  py -u export_pay_settlement.py --last-days 3       # 3 ngày gần nhất
  py -u export_pay_settlement.py --month 2026-06     # cả tháng (backfill)
  py -u export_pay_settlement.py --from 2026-06-01 --to 2026-06-30
  py -u export_pay_settlement.py --last-days 2 --xlsx bc.xlsx   # nạp BQ + ghi xlsx
  py -u export_pay_settlement.py --month 2026-06 --no-bq --xlsx bc.xlsx  # chỉ xlsx, không BQ

Mặc định nạp thẳng BigQuery rhysman.fact_tiktok_pay_settlement (idempotent: xoá đúng
khoảng settlement_time rồi insert). Cần cookie.txt + openpyxl + google-cloud-bigquery.

CHẠY TRÊN GITHUB ACTIONS — cấp qua biến môi trường, không cần file secret trong repo:
  TIKTOK_COOKIE                = nguyên chuỗi cookie (1 dòng)
  GOOGLE_SERVICE_ACCOUNT_JSON  = nội dung file service-account .json
  (tùy chọn) BQ_DATASET / BQ_TABLE để trỏ bảng khác.
An toàn: nếu TikTok đổi/thêm/bớt cột so với EXPECTED_HEADERS -> DỪNG báo lỗi.
"""
import argparse, datetime, gzip, io, json, os, re, sys, time, urllib.parse, urllib.request

# ================== CẤU HÌNH (ưu tiên ENV, fallback về giá trị local cũ) ==================
# Chạy LOCAL: không set ENV nào -> hành vi y hệt bản gốc (đọc cookie.txt + key ở ổ E:).
# Chạy GITHUB ACTIONS: set ENV -> không cần file secret nằm trong repo.

# --- Cookie TikTok Seller ---
COOKIE_FILE   = os.getenv("COOKIE_FILE", "cookie.txt")  # đổi qua wrapper cho từng shop
TIKTOK_COOKIE = os.getenv("TIKTOK_COOKIE_RHYSMAN")              # có giá trị -> dùng thẳng, bỏ qua file

SELLER_ID = None             # tự đọc từ cookie lúc chạy (oec_seller_id_unified_seller_env)

# --- BigQuery đích ---
# 3 cách cấp credential, xét theo thứ tự ưu tiên:
#   1) GOOGLE_SERVICE_ACCOUNT_JSON = NỘI DUNG chuỗi JSON  (dùng chung secret với keocptiktok.py)
#   2) BQ_KEY / GOOGLE_APPLICATION_CREDENTIALS = ĐƯỜNG DẪN tới file .json
#   3) không set gì -> dùng path mặc định ở máy anh
BQ_KEY_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
BQ_KEY      = os.getenv("BQ_KEY") or os.getenv("GOOGLE_APPLICATION_CREDENTIALS") \
              or "E:/RYSH MAN/rhysman-data-warehouse-488306-8db2b940e56a.json"
BQ_DATASET  = os.getenv("BQ_DATASET", "rhysman")
BQ_TABLE    = os.getenv("BQ_TABLE", "fact_tiktok_pay_settlement")

# ============================ ĐỊNH NGHĨA CỘT XUẤT RA ============================
# Bảng 'Chi tiết đơn hàng' gốc 65 cột -> rút còn các cột dưới. Tiền là VND (số nguyên).
# "Bồi thường do vận chuyển-kho vận" và "Các khoản giảm trừ..." KHÔNG phải cột —
# chúng là LOẠI GIAO DỊCH (cột 'Loại giao dịch'), tiền lấy ở cột 'Số tiền điều chỉnh'.
ADJ_SHIP_LOGISTICS = "Bồi thường do vận chuyển-kho vận"
ADJ_SELLER_DEDUCT  = "Các khoản giảm trừ do người bán chịu"

# 'Các khoản điều chỉnh' = gom mọi loại điều chỉnh KHÁC (catch-all) ngoài 2 loại trên.
SUM_TAX = ["Thuế GTGT do TikTok Shop khấu trừ", "Thuế TNCN do TikTok Shop khấu trừ",
           "Thuế bán hàng của voucher GMV Max"]
SUM_SHIPPING = ["Phí vận chuyển thực tế", "Phí vận chuyển trả hàng thực tế",
                "Phí vận chuyển của khách hàng được hoàn lại"]
SUM_SHIP_DISCOUNT = ["Chiết khấu phí vận chuyển của nền tảng", "Chi phí vận chuyển của khách hàng",
                     "Hoàn phí SFR", "Trợ cấp giao hàng không thành công", "Trợ giá vận chuyển"]
SUM_OTHER_SVC = ["Phí dịch vụ SFP", "Phí dịch vụ hoàn tiền thưởng",
                 "Phí dịch vụ Ưu đãi đặc biệt trên LIVE", "Phí dịch vụ Chương trình EAMS",
                 "Phí dịch vụ Flash Sale", "Phí chương trình TikTok PayLater",
                 "Phí nguồn lực chiến dịch", "Phí dịch vụ SFR", "Voucher GMV Max",
                 "Gói dịch vụ được quản lý (thuế bán hàng)", "Gói dịch vụ được quản lý (phí mỗi đơn hàng)",
                 "Phí quảng cáo GMV Max", "Tiền cọc hoa hồng liên kết", "Hoàn hoa hồng liên kết"]

# Thứ tự + cách lấy từng cột xuất ra. kind: col=lấy thẳng, const, sum=cộng nhiều cột,
# adj=khoản điều chỉnh đúng 1 loại, adjbucket=mọi loại điều chỉnh còn lại.
OUT_SPEC = [
    ("order_id",                                "col",  "ID đơn hàng/điều chỉnh"),
    ("related_order_id",                        "col",  "ID đơn hàng liên quan"),  # đơn gốc — để view GROUP BY gộp net theo đơn
    ("order_created_time",                      "col",  "Thời gian tạo đơn hàng"),
    ("settlement_time",                         "col",  "Thời gian quyết toán đơn hàng"),
    ("seller_id",                               "const", "__SELLER_ID__"),  # resolve lúc chạy
    ("subtotal_after_seller_discounts",         "col",  "Tổng phụ sau giảm giá của người bán"),
    ("refund_subtotal_after_seller_discounts",  "col",  "Tổng phụ của khoản hoàn tiền sau giảm giá của người bán"),
    ("transaction_fee",                         "col",  "Phí giao dịch"),
    ("tiktok_shop_commission_fee",              "col",  "Phí hoa hồng của TikTok Shop"),
    ("affiliate_commission",                    "col",  "Hoa hồng liên kết"),
    ("affiliate_shop_ads_commission",           "col",  "Hoa hồng liên kết Quảng cáo cửa hàng"),
    ("affiliate_partner_commission",            "col",  "Hoa hồng của đối tác liên kết"),
    ("affiliate_partner_shop_ads_commission",   "col",  "Hoa hồng quảng cáo cửa hàng của Đối tác liên kết"),
    ("voucher_xtra_service_fee",                "col",  "Phí dịch vụ Voucher Xtra"),
    ("order_processing_fee",                    "col",  "Phí xử lý đơn hàng"),
    ("shipping_logistics_compensation",         "adj",  ADJ_SHIP_LOGISTICS),
    ("seller_borne_deductions",                 "adj",  ADJ_SELLER_DEDUCT),
    ("withholding_tax",                         "sum",  SUM_TAX),
    ("shipping_fee",                            "sum",  SUM_SHIPPING),
    ("shipping_fee_discount",                   "sum",  SUM_SHIP_DISCOUNT),
    ("other_service_fees",                      "sum",  SUM_OTHER_SVC),
    ("other_adjustments",                       "adjbucket", None),
]

# 65 cột chuẩn đã biết của sheet 'Chi tiết đơn hàng' — để phát hiện TikTok đổi/thêm cột.
EXPECTED_HEADERS = [
    "ID đơn hàng/điều chỉnh", "Loại giao dịch", "Thời gian tạo đơn hàng",
    "Thời gian quyết toán đơn hàng", "Đơn vị tiền tệ", "Tổng số tiền quyết toán",
    "Tổng doanh thu", "Tổng phụ sau giảm giá của người bán", "Tổng phụ trước giảm giá",
    "Giảm giá của người bán", "Tổng phụ của khoản hoàn tiền sau giảm giá của người bán",
    "Tổng phụ hoàn tiền trước giảm giá của người bán", "Khoản hoàn tiền giảm giá của người bán",
    "Tổng phí", "Phí giao dịch", "Phí hoa hồng của TikTok Shop", "Phí vận chuyển của người bán",
    "Phí vận chuyển thực tế", "Chiết khấu phí vận chuyển của nền tảng",
    "Chi phí vận chuyển của khách hàng", "Phí vận chuyển trả hàng thực tế",
    "Phí vận chuyển của khách hàng được hoàn lại", "Hoàn phí SFR",
    "Trợ cấp giao hàng không thành công", "Trợ giá vận chuyển", "Hoa hồng liên kết",
    "Hoa hồng liên kết trước thuế TNCN (thuế thu nhập cá nhân)", "Thuế TNCN đã khấu trừ",
    "Hoa hồng liên kết Quảng cáo cửa hàng", "Hoa hồng của Quảng cáo cửa hàng liên kết trước thuế TNCN",
    "Thuế thu nhập cá nhân đã được khấu trừ vào khoản hoa hồng của Quảng cáo cửa hàng liên kết",
    "Hoa hồng của đối tác liên kết", "Tiền cọc hoa hồng liên kết", "Hoàn hoa hồng liên kết",
    "Hoa hồng quảng cáo cửa hàng của Đối tác liên kết", "Phí dịch vụ SFP",
    "Phí dịch vụ hoàn tiền thưởng", "Phí dịch vụ Ưu đãi đặc biệt trên LIVE",
    "Phí dịch vụ Voucher Xtra", "Phí xử lý đơn hàng", "Phí dịch vụ Chương trình EAMS",
    "Phí dịch vụ Flash Sale", "Thuế GTGT do TikTok Shop khấu trừ", "Thuế TNCN do TikTok Shop khấu trừ",
    "Phí chương trình TikTok PayLater", "Phí nguồn lực chiến dịch", "Phí dịch vụ SFR",
    "Voucher GMV Max", "Thuế bán hàng của voucher GMV Max", "Gói dịch vụ được quản lý (thuế bán hàng)",
    "Gói dịch vụ được quản lý (phí mỗi đơn hàng)", "Phí quảng cáo GMV Max", "Số tiền điều chỉnh",
    "ID đơn hàng liên quan", "Khách thanh toán", "Tiền hoàn của khách",
    "Giảm giá voucher đồng chi trả của người bán",
    "Hoàn tiền giảm giá voucher đồng chi trả của người bán", "Giảm giá của nền tảng",
    "Hoàn tiền giảm giá của nền tảng", "Giảm giá voucher đồng chi trả của nền tảng",
    "Hoàn tiền giảm giá voucher đồng chi trả của nền tảng", "Giảm phí vận chuyển của người bán",
    "Trọng lượng kiện hàng ước tính", "Trọng lượng kiện hàng được tính phí",
]

TZ = datetime.timezone(datetime.timedelta(hours=7))
BASE = "https://seller-vn.tiktok.com"
API = BASE + "/api/v2/pay/settlement/file"
def common_params():   # dựng theo SELLER_ID hiện tại (đã resolve từ cookie)
    return {"locale": "vi-VN", "language": "vi-VN", "oec_seller_id": SELLER_ID, "seller_id": SELLER_ID,
            "aid": "4068", "app_name": "i18n_ecom_shop", "device_platform": "web", "cookie_enabled": "true"}
HEADERS = {
    "user-agent": "Mozilla/5.0", "accept": "*/*",
    "origin": BASE, "referer": BASE + "/finance/transactions?shop_region=VN&tab=settled_tab",
    "x-tt-oec-region": "VN",
}
TIME_TYPE = {"settle": 1, "placed": 2}   # 1 = ngày quyết toán (xác nhận); 2 = ngày tạo đơn (suy đoán)

def req(url, cookie, body=None, raw=False, tries=5):
    for k in range(tries):
        try:
            data = json.dumps(body).encode() if body is not None else None
            h = {**HEADERS, "cookie": cookie}
            if body is not None:
                h["content-type"] = "application/json"
            r = urllib.request.Request(url, data=data, method="POST" if body is not None else "GET", headers=h)
            with urllib.request.urlopen(r, timeout=60) as resp:
                buf = resp.read()
                if resp.headers.get("content-encoding") == "gzip":
                    buf = gzip.decompress(buf)
            return buf if raw else json.loads(buf)
        except Exception as e:
            if k == tries - 1:
                raise
            print(f"    retry {k+1} ({e})", flush=True); time.sleep(2 * (k + 1))

def list_files(cookie):
    url = f"{API}/list?" + urllib.parse.urlencode({**common_params(), "page": 1, "page_size": 20, "size": 20, "from": 0})
    return (req(url, cookie).get("data") or {}).get("files", [])

# Cột giữ nguyên dạng chữ (ID, ngày); còn lại quy về số.
TEXT_COLS = {"ID đơn hàng/điều chỉnh", "ID đơn hàng liên quan", "Thời gian tạo đơn hàng", "Thời gian quyết toán đơn hàng"}

def _num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else v
    s = str(v).strip().replace(" ", "").replace(",", "")
    if not s or s == "-":
        return 0
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return 0

OUT_NAMES = [name for name, _, _ in OUT_SPEC]
DATE_OUT = {"order_created_time", "settlement_time"}   # cột output kiểu DATE

def _to_date(v):
    """'2026/06/25' -> '2026-06-25' (chuỗi ISO cho BigQuery DATE); rỗng -> None."""
    if v is None:
        return None
    s = str(v).strip().replace("/", "-")
    return s or None

def slim_records(src):
    """Đọc 'Chi tiết đơn hàng' (65 cột) -> list dict 21 cột (đã gộp + routing điều chỉnh).
    Ngày -> chuỗi ISO, tiền -> int, id -> str. src = đường dẫn hoặc bytes (xử lý trong RAM)."""
    import openpyxl
    if isinstance(src, (bytes, bytearray)):
        src = io.BytesIO(src)
    # KHÔNG dùng read_only: chế độ đó đọc dimension bị cắt cụt còn 21 cột (thiếu cột 21-64).
    wb = openpyxl.load_workbook(src, read_only=False, data_only=True)
    ws = wb["Chi tiết đơn hàng"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    need = ["Loại giao dịch", "Số tiền điều chỉnh"]
    for spec in OUT_SPEC:
        if spec[1] == "col":  need.append(spec[2])
        elif spec[1] == "sum": need += spec[2]
    # ---- LÁ CHẮN: cấu trúc file PHẢI khớp 65 cột đã biết, lệch là DỪNG ----
    actual = [str(h).strip() for h in header if h is not None]
    removed = [h for h in EXPECTED_HEADERS if h not in idx]
    added = [h for h in actual if h not in EXPECTED_HEADERS]
    missing = [c for c in need if c not in idx]
    if missing or removed or added:
        msg = ["[SLIM] CẤU TRÚC FILE THAY ĐỔI — DỪNG để bro kiểm tra trước khi tin số liệu:"]
        if missing: msg.append(f"  • MẤT/ĐỔI TÊN cột ĐANG DÙNG: {missing}")
        if removed: msg.append(f"  • Mất/đổi tên cột chuẩn: {removed}")
        if added:   msg.append(f"  • Cột MỚI xuất hiện: {added}")
        msg.append("  -> Cập nhật EXPECTED_HEADERS / OUT_SPEC cho khớp rồi chạy lại.")
        sys.exit("\n".join(msg))

    i_type = idx["Loại giao dịch"]; i_adj = idx["Số tiền điều chỉnh"]; i_id = idx["ID đơn hàng/điều chỉnh"]
    # plan: (out_name, kind, payload). kind: text/date/num/const/sum/adj/adjbucket
    plan = []
    for name, kind, src_col in OUT_SPEC:
        if kind == "col":
            k = "date" if name in DATE_OUT else ("text" if src_col in TEXT_COLS else "num")
            plan.append((name, k, idx[src_col]))
        elif kind == "const":     plan.append((name, "const", src_col))
        elif kind == "sum":       plan.append((name, "sum", [idx[c] for c in src_col]))
        elif kind == "adj":       plan.append((name, "adj", src_col))
        elif kind == "adjbucket": plan.append((name, "adjbucket", None))

    records, n_order, n_adj = [], 0, 0
    for row in it:
        if row is None or row[i_id] in (None, ""):
            continue
        ttype = row[i_type]; adj_amt = _num(row[i_adj])
        n_order += (ttype == "Đơn hàng"); n_adj += (ttype != "Đơn hàng")
        rec = {}
        for name, k, p in plan:
            if k == "num":         rec[name] = _num(row[p])
            elif k == "date":      rec[name] = _to_date(row[p])
            elif k == "text":      rec[name] = (str(row[p]) if row[p] not in (None, "") else None)
            elif k == "const":     rec[name] = SELLER_ID if p == "__SELLER_ID__" else p
            elif k == "sum":       rec[name] = sum(_num(row[c]) for c in p)
            elif k == "adj":       rec[name] = adj_amt if ttype == p else 0
            else:                  rec[name] = adj_amt if ttype not in ("Đơn hàng", ADJ_SHIP_LOGISTICS, ADJ_SELLER_DEDUCT) else 0
        records.append(rec)
    print(f"[SLIM] {len(records)} dòng ({n_order} đơn, {n_adj} điều chỉnh) x {len(OUT_SPEC)} cột", flush=True)
    return records

def write_records_xlsx(records, out_path):
    import openpyxl
    wb = openpyxl.Workbook(write_only=True); ws = wb.create_sheet("Chi tiết đơn hàng")
    ws.append(OUT_NAMES)
    for r in records:
        ws.append([r[n] for n in OUT_NAMES])
    wb.save(out_path)
    print(f"[XLSX] -> {out_path}", flush=True)

def slim_xlsx(src, out_path):      # tiện ích cũ: đọc + ghi xlsx
    write_records_xlsx(slim_records(src), out_path)

def _bq_credentials():
    """Dựng credential BigQuery. Ưu tiên chuỗi JSON trong ENV (hợp GitHub Actions),
    fallback về file .json trên đĩa (hợp chạy local)."""
    from google.oauth2 import service_account
    if BQ_KEY_JSON:
        try:
            info = json.loads(BQ_KEY_JSON)
        except json.JSONDecodeError as e:
            sys.exit(f"[BQ] GOOGLE_SERVICE_ACCOUNT_JSON không phải JSON hợp lệ: {e}")
        print("[BQ] credential: ENV GOOGLE_SERVICE_ACCOUNT_JSON", flush=True)
        return service_account.Credentials.from_service_account_info(info)
    if not os.path.exists(BQ_KEY):
        sys.exit(f"[BQ] không thấy key file: {BQ_KEY} — và cũng không có ENV GOOGLE_SERVICE_ACCOUNT_JSON.")
    print(f"[BQ] credential: file {BQ_KEY}", flush=True)
    return service_account.Credentials.from_service_account_file(BQ_KEY)

def load_to_bq(records, begin_date, end_date):
    """Nạp records vào {BQ_DATASET}.{BQ_TABLE}. Idempotent: xoá đúng khoảng
    settlement_time [begin_date, end_date] rồi insert lại."""
    from google.cloud import bigquery
    creds = _bq_credentials()
    client = bigquery.Client(credentials=creds, project=creds.project_id)
    tbl = f"{creds.project_id}.{BQ_DATASET}.{BQ_TABLE}"

    # 1) xoá dữ liệu cũ CỦA ĐÚNG SHOP NÀY trong khoảng (chạy lại không trùng;
    #    có seller_id để không xoá nhầm dữ liệu shop khác dùng chung bảng).
    q = (f"DELETE FROM `{tbl}` "
         f"WHERE seller_id = @sid AND settlement_time BETWEEN @b AND @e")
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("sid", "STRING", SELLER_ID),
        bigquery.ScalarQueryParameter("b", "DATE", begin_date),
        bigquery.ScalarQueryParameter("e", "DATE", end_date)])
    client.query(q, job_config=cfg).result()
    print(f"[BQ] đã xoá dữ liệu cũ của seller {SELLER_ID} trong {begin_date}..{end_date}", flush=True)

    # 2) insert (load job, schema khai báo rõ -> đúng kiểu DATE/INT64)
    def _bqtype(name):
        return "STRING" if name in ("order_id", "related_order_id", "seller_id") else ("DATE" if name in DATE_OUT else "INT64")
    schema = [bigquery.SchemaField(n, _bqtype(n)) for n in OUT_NAMES]
    # ALLOW_FIELD_ADDITION: lần đầu có cột mới (related_order_id) BQ tự thêm vào bảng, khỏi ALTER tay.
    job_cfg = bigquery.LoadJobConfig(schema=schema, write_disposition="WRITE_APPEND",
                                     schema_update_options=["ALLOW_FIELD_ADDITION"])
    job = client.load_table_from_json(records, tbl, job_config=job_cfg)
    job.result()
    print(f"[BQ] đã nạp {len(records)} dòng -> {tbl}", flush=True)

def main():
    ap = argparse.ArgumentParser(description="TikTok Seller export downloader")
    ap.add_argument("--month", help="YYYY-MM")
    ap.add_argument("--from", dest="d_from", help="YYYY-MM-DD")
    ap.add_argument("--to", dest="d_to", help="YYYY-MM-DD (bao gồm cả ngày này)")
    ap.add_argument("--last-days", type=int, default=None,
                    help="kéo N ngày gần nhất theo settlement_time (mặc định 2 nếu không nêu mode nào)")
    ap.add_argument("--by", choices=["settle", "placed"], default="settle")
    ap.add_argument("--xlsx", help="(tùy chọn) cũng ghi ra file xlsx 21 cột với tên này")
    ap.add_argument("--no-bq", action="store_true", help="KHÔNG nạp BigQuery (chỉ chạy/ghi xlsx)")
    ap.add_argument("--keep-raw", action="store_true", help="giữ luôn file đầy đủ 65 cột")
    ap.add_argument("--wait", type=int, default=180, help="số giây tối đa chờ server sinh file")
    a = ap.parse_args()
    global SELLER_ID
    if TIKTOK_COOKIE:
        cookie, cookie_src = TIKTOK_COOKIE.strip(), "ENV:TIKTOK_COOKIE"
    else:
        if not os.path.exists(COOKIE_FILE):
            sys.exit(f"[CONFIG] không thấy file {COOKIE_FILE}, và cũng không có ENV TIKTOK_COOKIE.")
        cookie, cookie_src = open(COOKIE_FILE, encoding="utf-8").read().strip(), f"file:{COOKIE_FILE}"
    m = re.search(r"oec_seller_id_unified_seller_env=(\d+)", cookie)
    if not m:
        sys.exit(f"[CONFIG] không thấy oec_seller_id trong cookie ({cookie_src}) — cookie sai/thiếu/HẾT HẠN?")
    SELLER_ID = m.group(1)
    print(f"[CONFIG] seller_id={SELLER_ID} | cookie={cookie_src}", flush=True)

    if a.month:
        y, m = map(int, a.month.split("-"))
        begin = int(datetime.datetime(y, m, 1, tzinfo=TZ).timestamp())
        end = int(datetime.datetime(y + (m == 12), (m % 12) + 1, 1, tzinfo=TZ).timestamp()) - 1
        tag = a.month.replace("-", "_")
    elif a.d_from and a.d_to:
        begin = int(datetime.datetime.strptime(a.d_from, "%Y-%m-%d").replace(tzinfo=TZ).timestamp())
        end = int(datetime.datetime.strptime(a.d_to, "%Y-%m-%d").replace(tzinfo=TZ).timestamp()) + 86400 - 1
        tag = f"{a.d_from}_{a.d_to}"
    else:
        # MẶC ĐỊNH (hợp cron): N ngày gần nhất theo settlement_time, N=2 nếu không nêu.
        n = a.last_days or 2
        today = datetime.datetime.now(TZ).date()
        begin_d = today - datetime.timedelta(days=n - 1)
        begin = int(datetime.datetime.combine(begin_d, datetime.time.min, tzinfo=TZ).timestamp())
        end = int(datetime.datetime.combine(today, datetime.time.max, tzinfo=TZ).timestamp())
        tag = f"last{n}d_{today:%Y%m%d}"
    begin_date = datetime.datetime.fromtimestamp(begin, TZ).date().isoformat()
    end_date = datetime.datetime.fromtimestamp(end, TZ).date().isoformat()
    print(f"[EXPORT] kỳ {datetime.datetime.fromtimestamp(begin,TZ):%Y-%m-%d} -> "
          f"{datetime.datetime.fromtimestamp(end,TZ):%Y-%m-%d} | theo "
          f"{'ngày quyết toán' if a.by=='settle' else 'ngày tạo đơn'}", flush=True)

    # 1) chụp danh sách file hiện có (để nhận ra file MỚI sau khi tạo)
    before = {f["file_id"] for f in list_files(cookie)}

    # 2) tạo job export
    body = {"period": {"begin_date": str(begin), "end_date": str(end), "time_type": TIME_TYPE[a.by]},
            "file_type": 1, "statement_version": 0}
    d = req(f"{API}/export?" + urllib.parse.urlencode(common_params()), cookie, body=body)
    if d.get("code") != 0:
        sys.exit(f"[EXPORT] tạo job lỗi: {d.get('code')} {d.get('message')}")
    print("[EXPORT] đã tạo job, chờ server sinh file...", flush=True)

    # 3) poll tới khi có file MỚI status=2 (sẵn sàng)
    t0 = time.time(); fid = fname = None
    while time.time() - t0 < a.wait:
        time.sleep(3)
        for f in list_files(cookie):
            if f["file_id"] not in before and f.get("status") == 2:
                fid, fname = f["file_id"], f.get("file_name", "")
                break
        if fid:
            break
        print(f"    …đang xử lý ({int(time.time()-t0)}s)", flush=True)
    if not fid:
        sys.exit(f"[EXPORT] quá {a.wait}s chưa thấy file. Server có thể đang chậm — chạy lại.")
    print(f"[EXPORT] file sẵn sàng: {fname} (id {fid})", flush=True)

    # 4) lấy link tải + tải về
    url = (req(f"{API}/download?" + urllib.parse.urlencode({**common_params(), "file_id": fid}), cookie)
           .get("data") or {}).get("url")
    if not url:
        sys.exit("[EXPORT] không lấy được link tải.")
    blob = req(BASE + url, cookie, raw=True)
    print(f"[EXPORT] tải xong file đầy đủ: {len(blob)//1024} KB", flush=True)

    if a.keep_raw:                      # tùy chọn: giữ file 65 cột để đối chiếu
        raw_path = f"_full_{tag}.xlsx"
        with open(raw_path, "wb") as fh:
            fh.write(blob)
        print(f"[EXPORT] giữ file đầy đủ: {raw_path}", flush=True)

    # parse 1 lần trong RAM -> records 21 cột
    records = slim_records(blob)
    if a.xlsx:                          # tùy chọn: cũng ghi xlsx
        write_records_xlsx(records, a.xlsx)
    if a.no_bq:
        print("[BQ] bỏ qua nạp BigQuery (--no-bq).", flush=True)
    else:                              # MẶC ĐỊNH: nạp BigQuery
        load_to_bq(records, begin_date, end_date)

if __name__ == "__main__":
    main()
